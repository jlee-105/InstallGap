"""
data_source.py
Excel 파일에서 IQ Sched Status Rpt / SIRFIS / Overrides / LRP 데이터를 읽어
표준 DataFrame으로 반환한다.
나중에 DB나 API 소스로 교체할 때 이 파일만 수정하면 된다.
"""

import openpyxl
import pandas as pd
from pathlib import Path


EXCEL_PATH = Path(__file__).parent.parent.parent / "data" / "1272-76-22_TI_2026-W30.5_RevisionReport.xlsx"


def _load_sheet_as_df(wb, sheet_name: str, header_row: int = 1) -> pd.DataFrame:

    ws = wb[sheet_name]
    data = []
    headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all(v is None for v in row):
            continue
        data.append(row)
    return pd.DataFrame(data, columns=headers)


def load_iq_status(wb):
    
    ws = wb["IQ Sched Status Rpt"]
    from openpyxl.utils import column_index_from_string
    max_col = column_index_from_string("KE")  # A~KE까지 읽음
    headers = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]

    data = []
    for row_idx in range(2, ws.max_row + 1):
        row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, max_col + 1)]
        # 완전 빈 행 제거
        if all(v is None for v in row_vals):
            continue
        data.append(row_vals)

    df = pd.DataFrame(data, columns=headers)
    return df


def load_sirfis(wb) -> pd.DataFrame:
    """SIRFIS 시트를 DataFrame으로 반환. Row 1=Export date, Row 3=실제 헤더."""
    return _load_sheet_as_df(wb, "SIRFIS", header_row=3)


def load_overrides(wb) -> pd.DataFrame:
    """Overrides 시트 반환. Row 4=헤더, Row 5~=데이터. 스위치(C2/H2)는 Frontend 입력."""
    ws = wb["Overrides"]
    headers = [ws.cell(row=4, column=c).value for c in range(1, 6)]
    data = []
    for row_idx in range(5, ws.max_row + 1):
        row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, 6)]
        if row_vals[0] is None:
            continue
        data.append(row_vals)
    return pd.DataFrame(data, columns=headers)


def load_ma(wb) -> pd.DataFrame:
    """MA into Output 시트를 DataFrame으로 반환. Row 1=헤더."""
    return _load_sheet_as_df(wb, "MA into Output", header_row=1)


def load_lrp_table(lrp_csv: Path) -> pd.DataFrame:
    """LRP 계획 테이블을 CSV에서 읽어 반환. 컬럼: date, wspw"""
    return pd.read_csv(lrp_csv, parse_dates=["date"])


# Symbol 폰트 0xE0=→ 등 Excel에서 잘못 저장된 문자 치환 테이블
_SYMBOL_CHAR_MAP = {
    '\u00e0': '→',
    '\u00e1': '←',
    '\u00e2': '↑',
    '\u00e3': '↓',
    '\n':     '',   # Excel Alt+Enter 줄바꿈 → 공백
}


def _fix_encoding(df: pd.DataFrame, from_enc: str, to_enc: str) -> pd.DataFrame:
    """문자열 셀과 컬럼명의 인코딩 깨짐을 재변환으로 복구."""
    def _convert(x):
        if not isinstance(x, str):
            return x
        try:
            x = x.encode(from_enc).decode(to_enc)
        except (UnicodeEncodeError, UnicodeDecodeError):
            pass
        for bad, good in _SYMBOL_CHAR_MAP.items():
            x = x.replace(bad, good)
        return x
    df = df.map(_convert)
    df.columns = [_convert(c) for c in df.columns]
    return df


def load_all(
    excel_path: Path = EXCEL_PATH,
    from_enc: str = "latin-1",
    to_enc: str = "utf-8",
) -> dict:

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    IQSR_df     = _fix_encoding(load_iq_status(wb), from_enc, to_enc)
    SIRFIS_df   = _fix_encoding(load_sirfis(wb),    from_enc, to_enc)
    OVERRIDE_df = _fix_encoding(load_overrides(wb), from_enc, to_enc)
    MA_df       = _fix_encoding(load_ma(wb),        from_enc, to_enc)
    wb.close()

    return {
        "iq":        IQSR_df,
        "sirfis":    SIRFIS_df,
        "overrides": OVERRIDE_df,
        "ma":        MA_df,
    }



# data = load_all()

# print("=== IQ rows:", len(data["iq"]))
# print("    columns:", list(data["iq"].columns), "...")
# a = input()

# print("\n=== SIRFIS rows:", len(data["sirfis"]))
# print("    columns:", list(data["sirfis"].columns), "...")

# a = input()

# print("\n=== Override rows:", len(data["overrides"]))
# print("    columns:", list(data["overrides"].columns))

# print("\n=== MA rows:", len(data["ma"]))
# print("    columns:", list(data["ma"].columns), "...")
